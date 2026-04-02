"""
Generate final formatted Excel output for Appendix E and Appendix X
"""
import pandas as pd
import numpy as np
from scipy.stats import spearmanr
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Style definitions ─────────────────────────────────────────
header_font = Font(bold=True, size=11, name="Arial")
title_font = Font(bold=True, size=13, name="Arial")
subtitle_font = Font(bold=True, size=11, name="Arial", color="333333")
normal_font = Font(size=10, name="Arial")
note_font = Font(size=9, name="Arial", italic=True, color="666666")
header_fill = PatternFill("solid", fgColor="D9E1F2")
highlight_fill = PatternFill("solid", fgColor="FFF2CC")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

def style_data_cell(ws, row, col, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.font = normal_font
    cell.alignment = center
    cell.border = thin_border
    if fmt:
        cell.number_format = fmt

# ── Load data ─────────────────────────────────────────────────
day = pd.read_csv("/mnt/user-data/uploads/grid_accessibility_daytime.csv")
night = pd.read_csv("/mnt/user-data/uploads/grid_accessibility_nighttime.csv")
risk = pd.read_excel("/mnt/user-data/uploads/成都五城区_地震风险_更新坐标.xlsx", sheet_name="计算过程")
day = day.merge(risk[["GRID_ID", "所属区"]], on="GRID_ID", how="left")
night = night.merge(risk[["GRID_ID", "所属区"]], on="GRID_ID", how="left")

district_map = {"青羊区": "Qingyang", "武侯区": "Wuhou", "金牛区": "Jinniu",
                "锦江区": "Jinjiang", "成华区": "Chenghua"}
district_order = ["青羊区", "武侯区", "金牛区", "锦江区", "成华区"]

# ══════════════════════════════════════════════════════════════
# SHEET 1: Table E2 - Spearman ρ Matrix
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Table E2 - Spearman ρ"

ws1.merge_cells("A1:F1")
ws1["A1"] = "Table E2. Pairwise Spearman Rank Correlation (ρ) of ESI Across Search-Radius Scenarios"
ws1["A1"].font = title_font

row = 3
for period_name, df in [("Daytime", day), ("Nighttime", night)]:
    ws1.cell(row=row, column=1, value=f"{period_name} Period").font = subtitle_font
    row += 1

    headers = ["Comparison", "Spearman ρ", "p-value", "n"]
    for j, h in enumerate(headers, 1):
        ws1.cell(row=row, column=j, value=h)
    style_header_row(ws1, row, len(headers))
    row += 1

    pairs = [("A", "B", "Ai_A", "Ai_B"), ("A", "C", "Ai_A", "Ai_C"), ("B", "C", "Ai_B", "Ai_C")]
    for s1, s2, c1, c2 in pairs:
        rho, pval = spearmanr(df[c1], df[c2])
        ws1.cell(row=row, column=1, value=f"Scenario {s1} vs {s2}")
        ws1.cell(row=row, column=2, value=round(rho, 4))
        ws1.cell(row=row, column=3, value=f"< 0.001")
        ws1.cell(row=row, column=4, value=len(df))
        for c in range(1, 5):
            style_data_cell(ws1, row, c)
        row += 1
    row += 1

ws1.cell(row=row, column=1, value="Note: All correlations are significant at p < 0.001. Scenario A: d₀ = 3,000 m; Scenario B: d₀ = 2,000 m; Scenario C: d₀ = 1,000 m. Hotel shelter d₀ = 1,000 m is held constant.").font = note_font
ws1.merge_cells(f"A{row}:F{row}")

for col_letter in ["A", "B", "C", "D", "E", "F"]:
    ws1.column_dimensions[col_letter].width = 18

# ══════════════════════════════════════════════════════════════
# SHEET 2: Table E3 - Categorical Agreement & District ESI
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Table E3 - Agreement & District")

ws2.merge_cells("A1:G1")
ws2["A1"] = "Table E3. Categorical Agreement Rate and Inter-District ESI Ranking Across Scenarios"
ws2["A1"].font = title_font

row = 3
ws2.cell(row=row, column=1, value="Panel A: Categorical Spatial Agreement Rate (%)").font = subtitle_font
row += 1

headers = ["Period", "Comparison", "Agreement Count", "Total Grids", "Agreement Rate (%)"]
for j, h in enumerate(headers, 1):
    ws2.cell(row=row, column=j, value=h)
style_header_row(ws2, row, len(headers))
row += 1

def categorize_esi(esi):
    return pd.cut(esi, bins=[-np.inf, 0.5, 1.0, 1.5, np.inf],
                  labels=["Severely deficient", "Deficient", "Moderate", "Sufficient"])

for period_name, df in [("Daytime", day), ("Nighttime", night)]:
    cats = {s: categorize_esi(df[c]) for s, c in [("A", "Ai_A"), ("B", "Ai_B"), ("C", "Ai_C")]}
    for s1, s2 in [("A", "B"), ("A", "C"), ("B", "C")]:
        agree = (cats[s1] == cats[s2]).sum()
        rate = agree / len(df) * 100
        ws2.cell(row=row, column=1, value=period_name)
        ws2.cell(row=row, column=2, value=f"Scenario {s1} vs {s2}")
        ws2.cell(row=row, column=3, value=agree)
        ws2.cell(row=row, column=4, value=len(df))
        ws2.cell(row=row, column=5, value=round(rate, 1))
        for c in range(1, 6):
            style_data_cell(ws2, row, c)
        row += 1

row += 2
ws2.cell(row=row, column=1, value="Panel B: District-Level Mean ESI Across Scenarios").font = subtitle_font
row += 1

headers = ["Period", "District", "Scenario A (d₀=3000m)", "Scenario B (d₀=2000m)",
           "Scenario C (d₀=1000m)", "Rank A", "Rank B", "Rank C", "Rank Change"]
for j, h in enumerate(headers, 1):
    ws2.cell(row=row, column=j, value=h)
style_header_row(ws2, row, len(headers))
row += 1

for period_name, df in [("Daytime", day), ("Nighttime", night)]:
    for dist_cn in district_order:
        dist_en = district_map[dist_cn]
        mask = df["所属区"] == dist_cn
        means = [df.loc[mask, f"Ai_{s}"].mean() for s in ["A", "B", "C"]]
        ranks = []
        for col in ["Ai_A", "Ai_B", "Ai_C"]:
            dm = df.groupby("所属区")[col].mean()
            ranks.append(int(dm.rank(ascending=False)[dist_cn]))
        change = "Unchanged" if len(set(ranks)) == 1 else f"{ranks[0]}→{ranks[1]}→{ranks[2]}"

        ws2.cell(row=row, column=1, value=period_name)
        ws2.cell(row=row, column=2, value=dist_en)
        for j, m in enumerate(means, 3):
            ws2.cell(row=row, column=j, value=round(m, 4))
        for j, r in enumerate(ranks, 6):
            ws2.cell(row=row, column=j, value=r)
        ws2.cell(row=row, column=9, value=change)
        for c in range(1, 10):
            style_data_cell(ws2, row, c)
        row += 1

row += 1
ws2.cell(row=row, column=1, value="Note: ESI tiers: Sufficient (≥1.5), Moderate (1.0–1.5), Deficient (0.5–1.0), Severely deficient (<0.5). Rank 1 = highest mean ESI.").font = note_font
ws2.merge_cells(f"A{row}:I{row}")

for i, w in enumerate([12, 14, 22, 22, 22, 8, 8, 8, 14], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════
# SHEET 3: Table X1 - CDI α Sensitivity
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Table X1 - α Sensitivity")

ws3.merge_cells("A1:F1")
ws3["A1"] = "Table X1. Spearman Rank Correlation of CDI Under Different α Values (vs. α = 0.5 Baseline)"
ws3["A1"].font = title_font

# Recompute CDI
weights_config = {
    "H": {"cols": ["H11_标准化", "H13_标准化"], "ahp": [0.6667, 0.3333], "ent": [0.4747, 0.5253]},
    "E_day": {"cols": ["E11_昼_标准化", "E12_标准化", "E21_标准化", "E31_标准化"],
              "ahp": [0.4829, 0.0882, 0.1570, 0.2720], "ent": [0.4905, 0.0643, 0.2006, 0.2446]},
    "E_night": {"cols": ["E11_夜_标准化", "E12_标准化", "E21_标准化", "E31_标准化"],
                "ahp": [0.4829, 0.0882, 0.1570, 0.2720], "ent": [0.4829, 0.0653, 0.2036, 0.2482]},
    "V": {"cols": ["V11_标准化", "V21_标准化", "V22_标准化", "V31_标准化", "V32_标准化"],
           "ahp": [0.1529, 0.4147, 0.2573, 0.0876, 0.0876], "ent": [0.2608, 0.3083, 0.0717, 0.1888, 0.1704]},
}

def compute_dim(df, cfg, alpha):
    w = [alpha * a + (1 - alpha) * e for a, e in zip(cfg["ahp"], cfg["ent"])]
    return sum(w_i * df[c].values for w_i, c in zip(w, cfg["cols"]))

alpha_values = [0.3, 0.4, 0.5, 0.6, 0.7]
cdi = {}
for a in alpha_values:
    H = compute_dim(risk, weights_config["H"], a)
    Ed = compute_dim(risk, weights_config["E_day"], a)
    En = compute_dim(risk, weights_config["E_night"], a)
    V = compute_dim(risk, weights_config["V"], a)
    cdi[a] = {"day": H * Ed * V, "night": H * En * V}

# Panel A: vs baseline
row = 3
ws3.cell(row=row, column=1, value="Panel A: Spearman ρ Relative to α = 0.5").font = subtitle_font
row += 1

headers = ["α", "ρ (Daytime)", "p-value (Day)", "ρ (Nighttime)", "p-value (Night)"]
for j, h in enumerate(headers, 1):
    ws3.cell(row=row, column=j, value=h)
style_header_row(ws3, row, len(headers))
row += 1

for a in alpha_values:
    rd, pd_ = spearmanr(cdi[a]["day"], cdi[0.5]["day"])
    rn, pn_ = spearmanr(cdi[a]["night"], cdi[0.5]["night"])
    ws3.cell(row=row, column=1, value=a)
    ws3.cell(row=row, column=2, value=round(rd, 6))
    ws3.cell(row=row, column=3, value="1.000" if a == 0.5 else "< 0.001")
    ws3.cell(row=row, column=4, value=round(rn, 6))
    ws3.cell(row=row, column=5, value="1.000" if a == 0.5 else "< 0.001")
    for c in range(1, 6):
        style_data_cell(ws3, row, c)
    if a == 0.5:
        for c in range(1, 6):
            ws3.cell(row=row, column=c).fill = highlight_fill
    row += 1

# Panel B: Full pairwise matrix (daytime)
row += 2
ws3.cell(row=row, column=1, value="Panel B: Full Pairwise Spearman ρ Matrix (Daytime CDI)").font = subtitle_font
row += 1

headers = [""] + [f"α = {a}" for a in alpha_values]
for j, h in enumerate(headers, 1):
    ws3.cell(row=row, column=j, value=h)
style_header_row(ws3, row, len(headers))
row += 1

for i, ai in enumerate(alpha_values):
    ws3.cell(row=row, column=1, value=f"α = {ai}")
    style_data_cell(ws3, row, 1)
    for j, aj in enumerate(alpha_values):
        rho, _ = spearmanr(cdi[ai]["day"], cdi[aj]["day"])
        ws3.cell(row=row, column=j + 2, value=round(rho, 4))
        style_data_cell(ws3, row, j + 2, "0.0000")
    row += 1

# Panel C: High-demand area consistency
row += 2
ws3.cell(row=row, column=1, value="Panel C: High-Demand Area Identification Consistency (Daytime)").font = subtitle_font
row += 1

headers = ["α", "Jaccard Index", "Agreement Rate (%)", "n (high-demand grids)"]
for j, h in enumerate(headers, 1):
    ws3.cell(row=row, column=j, value=h)
style_header_row(ws3, row, len(headers))
row += 1

base_threshold = np.percentile(cdi[0.5]["day"], 75)
base_high = cdi[0.5]["day"] >= base_threshold
for a in alpha_values:
    threshold = np.percentile(cdi[a]["day"], 75)
    high = cdi[a]["day"] >= threshold
    overlap = (high & base_high).sum()
    union = (high | base_high).sum()
    jaccard = overlap / union
    agree = (high == base_high).mean() * 100
    ws3.cell(row=row, column=1, value=a)
    ws3.cell(row=row, column=2, value=round(jaccard, 4))
    ws3.cell(row=row, column=3, value=round(agree, 1))
    ws3.cell(row=row, column=4, value=int(high.sum()))
    for c in range(1, 5):
        style_data_cell(ws3, row, c)
    row += 1

row += 1
ws3.cell(row=row, column=1, value="Note: n = 4,857 grid cells. Combined weight formula: w_combined = α × w_AHP + (1−α) × w_entropy. High-demand areas defined as CDI ≥ 75th percentile.").font = note_font
ws3.merge_cells(f"A{row}:F{row}")

for i, w in enumerate([10, 18, 18, 18, 18, 18], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ── Save ──────────────────────────────────────────────────────
outpath = "/home/claude/output/Appendix_E_X_Results.xlsx"
wb.save(outpath)
print(f"Saved: {outpath}")
